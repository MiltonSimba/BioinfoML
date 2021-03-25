import openpyxl as xl      #xl is shortening the process of always typing the whole  openpyxl word

from openpyxl.chart import BarChart, Reference      #from the openpyxl package we are importing the chart module with classes BarChart and reference


def process_workbook(filename)
    wb = xl.load_workbook('filename')     #wb is the object variable used to retain the rhd expression value
    sheet = wb['Sheet1']                         #case sensitive Sheet1 workboook [] acceses items in the excel
    #cell = sheet ['a1']                      #returns same value as below
    #cell = sheet.cell (1,1)

    #print(cell.value)         #prints the string in cell variable

    #print (sheet.max_row)              #calculates how many rows we've in a spreadsheet

    for row in range (2,sheet.max_row+1):       #the +1 is making sure the row 4 is  included in analysis
        cell = sheet.cell(row,3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell (row,4)
        corrected_price_cell.value = corrected_price

    #reference (sheet, min_row = 2, max_row=sheet.max_row,min_col=4,max_col)
    #can be simplified and be more readable in this setup

    values = Reference (sheet, 
                min_row = 2, 
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)


    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart (chart, 'e2')

    wb.save('filename.xlsx')
